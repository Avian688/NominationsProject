import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule


# ---- settings ----
INPUT_CSV = "Sussex Awards 2025_November 4, 2025_12.33.csv"
OUTPUT_XLSX = "sussex_awards_extracted.xlsx"

# New column (empty) that should NOT count toward completeness
NGA_COL_NAME = "Nominee's NGA number"


def first_sentence(t):
    if not isinstance(t, str):
        return ""
    for sep in [".", ":"]:
        if sep in t:
            return t.split(sep)[0].strip() + sep
    return t.strip()


def find_column_for_question(df, question_text):
    target = first_sentence(question_text)
    question_row = df.iloc[0]
    for col in df.columns:
        cell = question_row[col]
        if isinstance(cell, str) and first_sentence(cell).startswith(target):
            return col
    raise ValueError(f"Could not find a column for question text: {question_text!r}")


def norm_text(x):
    if pd.isna(x):
        return ""
    if not isinstance(x, str):
        x = str(x)
    return " ".join(x.strip().lower().split())


def is_nonempty(v) -> bool:
    if pd.isna(v):
        return False
    return bool(str(v).strip())


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    """
    Excel sheet name rules:
      - max 31 chars
      - cannot contain: : \\ / ? * [ ]
    """
    if name is None:
        name = ""
    s = str(name).strip()
    if not s:
        s = "Category"

    bad_chars = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in bad_chars:
        s = s.replace(ch, " ")

    s = " ".join(s.split()).strip()
    if not s:
        s = "Category"

    s = s[:31]

    base = s
    i = 2
    while s in used:
        suffix = f" {i}"
        s = (base[: 31 - len(suffix)] + suffix).strip()
        i += 1

    used.add(s)
    return s


def dataframe_incomplete_mask(df: pd.DataFrame, ignore_columns: set[str] | None = None) -> pd.Series:
    """
    A row is incomplete if ANY cell is empty/blank/NaN,
    EXCEPT columns listed in ignore_columns.
    """
    ignore_columns = ignore_columns or set()

    def cell_blank(x) -> bool:
        if pd.isna(x):
            return True
        return str(x).strip() == ""

    def row_incomplete(row) -> bool:
        for col, val in row.items():
            if col in ignore_columns:
                continue
            if cell_blank(val):
                return True
        return False

    return df.apply(row_incomplete, axis=1)


def main():
    df = pd.read_csv(INPUT_CSV)

    # Qualtrics export: row 0 = question text, row 1 = ImportId, data from row 2
    data = df.iloc[2:].reset_index(drop=True)

    # ---- locate key columns ----
    questions = {
        "first_name": "First name:",
        "surname": "Surname:",
        "sussex_email": "Your Sussex email address:",
        "staff_student": "Are you staff, student or a postgraduate researcher?",
        "school_division": "Select your School or Division:",
        "nominee_name": "Name of nominee/s:",
        "nominee_email": "Nominee's Sussex email address (if known):",
        "nominee_school": "Select School or Division of nominee (if known):",
        "category": "Before choosing a category, please review the criteria to ensure your nomination meets the requirements:",
    }
    cols = {k: find_column_for_question(df, v) for k, v in questions.items()}

    # Prefer literal Q6 if present (more robust)
    category_col = "Q6" if "Q6" in df.columns else cols["category"]

    # ---- merge all responses to the right of Q6 into Category description (responses only, no headers) ----
    q6_idx = list(df.columns).index(category_col)
    cols_after_q6 = list(df.columns)[q6_idx + 1 :]

    def merged_category_description(row) -> str:
        parts = []
        for c in cols_after_q6:
            v = row[c]
            if is_nonempty(v):
                parts.append(str(v).strip())
        return "\n\n".join(parts).strip()

    # ---- build output dataframe (column order here) ----
    out_df = pd.DataFrame()
    out_df["First name:"] = data[cols["first_name"]]
    out_df["Surname:"] = data[cols["surname"]]
    out_df["Your Sussex Email address:"] = data[cols["sussex_email"]]
    out_df["are you staff, student or a postgraduate"] = data[cols["staff_student"]]
    out_df["select your school or division"] = data[cols["school_division"]]

    out_df["Self nomination?"] = ""

    # New empty NGA column (must be left of nominee name)
    out_df[NGA_COL_NAME] = ""

    out_df["Name of nominee(s)"] = data[cols["nominee_name"]]
    out_df["Nominee's Sussex email address (if known):"] = data[cols["nominee_email"]]
    out_df["Select School or Division of nominee (if known):"] = data[cols["nominee_school"]]

    out_df["Category"] = data[category_col]
    out_df["Category description"] = data.apply(merged_category_description, axis=1)

    # ---- self nomination logic ----
    for idx in range(len(out_df)):
        first = norm_text(data.at[idx, cols["first_name"]])
        surname = norm_text(data.at[idx, cols["surname"]])
        full_name = (first + " " + surname).strip()

        nom_name = norm_text(data.at[idx, cols["nominee_name"]])
        email = norm_text(data.at[idx, cols["sussex_email"]])
        nom_email = norm_text(data.at[idx, cols["nominee_email"]])

        is_self = False
        if full_name and nom_name and full_name == nom_name:
            is_self = True
        if email and nom_email and email == nom_email:
            is_self = True

        out_df.at[idx, "Self nomination?"] = "Yes" if is_self else "No"

    # ---- split by category, move incomplete rows to Incomplete ----
    # IMPORTANT: NGA column is allowed to be empty, so ignore it for completeness checks.
    ignore_for_completeness = {NGA_COL_NAME}

    incomplete_mask_all = dataframe_incomplete_mask(out_df, ignore_columns=ignore_for_completeness)
    incomplete_df = out_df.loc[incomplete_mask_all].copy()
    complete_df = out_df.loc[~incomplete_mask_all].copy()

    # Unique categories from complete rows only (future-proof)
    categories = (
        complete_df["Category"]
        .astype(str)
        .map(lambda x: x.strip())
        .loc[lambda s: s != ""]
        .unique()
        .tolist()
    )

    # Write: Incomplete first, then one sheet per category (no All, no Uncategorized)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        incomplete_df.to_excel(writer, sheet_name="Incomplete", index=False)

        used_names = {"Incomplete"}
        for cat in categories:
            sheet_name = sanitize_sheet_name(cat, used_names)
            subset = complete_df[complete_df["Category"].astype(str).str.strip() == cat].copy()
            subset.to_excel(writer, sheet_name=sheet_name, index=False)

    # ---- style workbook: filters, wrap text, keep rows small, conditional pale-red blanks on Incomplete ----
    wb = load_workbook(OUTPUT_XLSX)

    pale_red_fill = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
    header_font = Font(bold=True)

    # Keep your existing compact layout (fixed heights)
    header_height = 15
    row_height = 15
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        # Filters
        if max_row >= 1 and max_col >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # Freeze header
        ws.freeze_panes = "A2"

        # Header style
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.alignment = wrap_align

        ws.row_dimensions[1].height = header_height

        # Data cells: wrap + fixed small row height
        for r in range(2, max_row + 1):
            ws.row_dimensions[r].height = row_height
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).alignment = wrap_align

        # Column widths (reasonable defaults; make description wider)
        headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        for c in range(1, max_col + 1):
            h = headers[c - 1] if isinstance(headers[c - 1], str) else ""
            if h.strip().lower() == "category description":
                ws.column_dimensions[get_column_letter(c)].width = 60
            elif "email" in h.lower():
                ws.column_dimensions[get_column_letter(c)].width = 28
            elif h.strip() == NGA_COL_NAME:
                ws.column_dimensions[get_column_letter(c)].width = 18
            else:
                ws.column_dimensions[get_column_letter(c)].width = 22

    # Conditional formatting on Incomplete: highlight blank cells (stays highlighted until filled)
    # BUT: do NOT highlight NGA blanks (since it's allowed to be empty)
    if "Incomplete" in wb.sheetnames:
        ws_inc = wb["Incomplete"]
        max_row = ws_inc.max_row
        max_col = ws_inc.max_column

        if max_row >= 2 and max_col >= 1:
            # Identify NGA column index (if present)
            nga_col_idx = None
            for c in range(1, max_col + 1):
                if ws_inc.cell(row=1, column=c).value == NGA_COL_NAME:
                    nga_col_idx = c
                    break

            for c in range(1, max_col + 1):
                if nga_col_idx is not None and c == nga_col_idx:
                    continue  # skip NGA column

                col_letter = get_column_letter(c)
                formula = f'LEN(TRIM({col_letter}2))=0'
                rule = FormulaRule(formula=[formula], fill=pale_red_fill)
                ws_inc.conditional_formatting.add(f"{col_letter}2:{col_letter}{max_row}", rule)

    wb.save(OUTPUT_XLSX)
    print(f"Done! Written to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
